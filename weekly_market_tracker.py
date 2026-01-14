#!/usr/bin/env python3
"""Weekly Market Analysis Tracker"""

import json
import os
import pickle
import sys
import time
import warnings
from datetime import datetime, timedelta
from pathlib import Path

import numpy as np
import pandas as pd
import requests
import yfinance as yf
from dotenv import load_dotenv
from ta.trend import ADXIndicator
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, Rule
from openpyxl.styles import PatternFill
from openpyxl.styles.differential import DifferentialStyle

warnings.filterwarnings('ignore')

# Load environment variables from .env file
load_dotenv()

# CONFIGURATION

FRED_API_KEY = os.getenv("FRED_API_KEY")
if not FRED_API_KEY:
    raise ValueError("FRED_API_KEY environment variable is not set. Please check your .env file.")

# Default assets file
SCRIPT_DIR = Path(__file__).parent
DEFAULT_ASSETS_FILE = SCRIPT_DIR / "assets.json"


def load_assets(file_path=None):
    """Load assets from JSON file."""
    if file_path is None:
        file_path = DEFAULT_ASSETS_FILE
    else:
        file_path = Path(file_path)

    if not file_path.exists():
        print(f"Error: Assets file not found: {file_path}")
        sys.exit(1)

    with open(file_path, 'r') as f:
        return json.load(f)

MA_PERIODS = [20, 50, 100, 200]
ZSCORE_WINDOW = 20

# CACHE CONFIGURATION
CACHE_DIR = SCRIPT_DIR / ".cache"
CACHE_DIR.mkdir(exist_ok=True)


def get_cache_path(cache_key):
    """Get cache file path for today."""
    today = datetime.now().strftime('%Y%m%d')
    return CACHE_DIR / f"{cache_key}_{today}.json"


def load_from_cache(cache_key):
    """Load data from cache if exists and is from today."""
    cache_file = get_cache_path(cache_key)
    if cache_file.exists():
        try:
            with open(cache_file, 'r') as f:
                return json.load(f)
        except (json.JSONDecodeError, IOError):
            pass
    return None


def save_to_cache(cache_key, data):
    """Save data to cache with today's date."""
    cache_file = get_cache_path(cache_key)
    # Clean old cache files for this key
    for old_file in CACHE_DIR.glob(f"{cache_key}_*.json"):
        if old_file != cache_file:
            old_file.unlink()
    with open(cache_file, 'w') as f:
        json.dump(data, f)


def cleanup_orphan_caches(max_age_days=7):
    """Remove orphan cache files older than max_age_days."""
    cutoff_time = time.time() - (max_age_days * 86400)
    for cache_file in CACHE_DIR.glob("*_historical.pkl"):
        if cache_file.stat().st_mtime < cutoff_time:
            cache_file.unlink()


def get_cached_ticker(ticker, period, interval):
    """
    Smart persistent cache for yfinance data.
    Returns cached data if last bar is recent, otherwise fetches fresh data.

    Best practices:
    - Uses auto_adjust=True for split/dividend adjusted prices
    - Handles timezone-aware datetimes properly
    - Validates data before caching
    """
    cache_file = CACHE_DIR / f"{ticker}_{interval}_historical.pkl"

    # Load existing cache
    if cache_file.exists():
        try:
            cached_data = pd.read_pickle(cache_file)
            if not cached_data.empty and len(cached_data) > 0:
                # Convert timezone-aware to date for comparison
                last_date = cached_data.index[-1]
                if hasattr(last_date, 'date'):
                    last_date = last_date.date()
                else:
                    last_date = pd.Timestamp(last_date).date()

                # If last bar is today or yesterday, use cache
                yesterday = (datetime.now() - timedelta(days=1)).date()
                if last_date >= yesterday:
                    return cached_data
        except (FileNotFoundError, EOFError, pd.errors.EmptyDataError, pickle.PickleError):
            pass

    # Fetch fresh data with best practices
    try:
        data = yf.download(
            ticker,
            period=period,
            interval=interval,
            auto_adjust=True,  # Adjust for splits/dividends
            progress=False,
            ignore_tz=False  # Preserve timezone info
        )

        # Handle MultiIndex columns (happens with single ticker sometimes)
        if isinstance(data.columns, pd.MultiIndex):
            data.columns = data.columns.get_level_values(0)

        # Validate data before caching
        if not data.empty and len(data) > 0 and 'Close' in data.columns:
            data.to_pickle(cache_file)
            return data
        else:
            # Return empty DataFrame if download failed
            return pd.DataFrame()

    except Exception as e:
        print(f"Warning: Failed to download {ticker}: {e}")
        return pd.DataFrame()


# DATA FETCHING

def get_fred_series(series_id, limit=1):
    """Fetch values from FRED API."""
    params = {
        "series_id": series_id,
        "api_key": FRED_API_KEY,
        "file_type": "json",
        "sort_order": "desc",
        "limit": limit,
    }
    r = requests.get("https://api.stlouisfed.org/fred/series/observations", params=params, timeout=10)
    r.raise_for_status()
    
    observations = r.json().get('observations', [])
    if not observations:
        return (None, None) if limit == 1 else []
    
    if limit == 1:
        return float(observations[0]['value']), observations[0]['date']
    
    return [(float(o['value']), o['date']) for o in observations if o['value'] != '.']


def get_fear_greed_traditional():
    """Fetch CNN Fear & Greed Index using fear-and-greed package."""
    cache_key = "fear_greed_stocks"
    cached = load_from_cache(cache_key)
    if cached:
        return cached['value'], cached['description']

    try:
        import fear_and_greed
    except ImportError:
        raise ImportError(
            "fear-and-greed package not installed.\n"
            "Install with: pip install fear-and-greed"
        )

    data = fear_and_greed.get()
    result = {'value': round(data.value), 'description': data.description.title()}
    save_to_cache(cache_key, result)
    return result['value'], result['description']


def get_fear_greed_crypto():
    """Fetch Crypto Fear & Greed Index."""
    cache_key = "fear_greed_crypto"
    cached = load_from_cache(cache_key)
    if cached:
        return cached['value'], cached['classification']

    r = requests.get("https://api.alternative.me/fng/?limit=1", timeout=10)
    r.raise_for_status()
    data = r.json()['data'][0]
    result = {'value': int(data['value']), 'classification': data['value_classification'].title()}
    save_to_cache(cache_key, result)
    return result['value'], result['classification']


def get_vix_zscore():
    """Fetch VIX and calculate smoothed inverted Z-score (252-day rolling window, 5-period EMA)."""
    cache_key = "vix_zscore"
    cached = load_from_cache(cache_key)
    if cached:
        return cached['vix'], cached['z_score']

    hist = get_cached_ticker("^VIX", period="2y", interval="1d")
    if hist.empty or len(hist) < 252:
        return None, None

    vix_series = hist['Close']
    vix = round(vix_series.iloc[-1], 2)

    # Z-score with 252-day window (1 year of trading days)
    mean = vix_series.rolling(252).mean()
    std = vix_series.rolling(252).std()
    z = (vix_series - mean) / std

    # Invert and smooth with 5-period EMA
    z_inverted = -z
    z_smooth = z_inverted.ewm(span=5, adjust=False).mean()

    z_score = round(z_smooth.iloc[-1], 2)
    result = {'vix': vix, 'z_score': z_score}
    save_to_cache(cache_key, result)
    return vix, z_score


# GLI CALCULATION

def calculate_gli():
    """Calculate Global Liquidity Index: Fed Balance Sheet - TGA - RRP."""
    cache_key = "gli"
    cached = load_from_cache(cache_key)
    if cached:
        return cached

    fed_data = get_fred_series("WALCL", limit=14)
    tga_data = get_fred_series("WTREGEN", limit=70)
    rrp_data = get_fred_series("RRPONTSYD", limit=70)

    if not all([fed_data, tga_data, rrp_data]):
        return None

    gli_series = []
    for fed_val, fed_date in fed_data:
        # Find most recent TGA and RRP values on or before fed_date
        tga_val = next((v for v, d in tga_data if d <= fed_date), None)
        rrp_val = next((v for v, d in rrp_data if d <= fed_date), None)
        if tga_val and rrp_val:
            gli_series.append(((fed_val - tga_val - rrp_val) / 1000, fed_date))

    if not gli_series:
        return None

    current_gli, current_date = gli_series[0]

    def calc_change(weeks):
        if len(gli_series) > weeks:
            prev = gli_series[weeks][0]
            change = current_gli - prev
            return round(change, 2), round((change / prev) * 100, 2)
        return None, None

    wow_change, wow_pct = calc_change(1)
    mom_change, mom_pct = calc_change(4)
    qoq_change, qoq_pct = calc_change(12)

    trend = "📈 Expanding" if mom_pct and mom_pct > 1 else "📉 Contracting" if mom_pct and mom_pct < -1 else "➡️ Flat"

    result = {
        'value': round(current_gli, 2),
        'fed_bs': round(fed_data[0][0] / 1000, 2),
        'tga': round(tga_data[0][0] / 1000, 2),
        'rrp': round(rrp_data[0][0] / 1000, 2),
        'date': current_date,
        'wow_change': wow_change, 'wow_pct': wow_pct,
        'mom_change': mom_change, 'mom_pct': mom_pct,
        'qoq_change': qoq_change, 'qoq_pct': qoq_pct,
        'trend': trend,
    }
    save_to_cache(cache_key, result)
    return result


# TECHNICAL ANALYSIS

def calculate_tema(series, period):
    """
    Calculate Triple Exponential Moving Average (TEMA).
    TEMA = 3*EMA - 3*EMA(EMA) + EMA(EMA(EMA))
    More responsive than EMA, less lag.
    """
    ema1 = series.ewm(span=period, adjust=False).mean()
    ema2 = ema1.ewm(span=period, adjust=False).mean()
    ema3 = ema2.ewm(span=period, adjust=False).mean()
    tema = 3 * ema1 - 3 * ema2 + ema3
    return tema


def get_volatility_scalar(close, lookback=252):
    """
    Calculate volatility scalar for adaptive period selection.
    Returns ratio of current ATR vs historical median ATR.

    High volatility (ratio > 1) → use longer periods to filter noise
    Low volatility (ratio < 1) → use shorter periods to capture moves

    Clamped to [0.5, 2.0] to prevent extreme period adjustments.
    """
    if len(close) < lookback:
        return 1.0  # Default to no adjustment if insufficient data

    # Calculate ATR (14-day rolling mean of absolute returns)
    atr = close.diff().abs().rolling(14).mean()

    # Current ATR vs historical median
    atr_current = atr.iloc[-1]
    atr_historical = atr.rolling(lookback).median().iloc[-1]

    if atr_historical == 0 or pd.isna(atr_historical):
        return 1.0

    ratio = atr_current / atr_historical

    # Clamp to prevent extreme values
    return np.clip(ratio, 0.5, 2.0)


def calculate_tema_ensemble(close, price):
    """
    Multi-period TEMA ensemble with volatility adjustment.

    Calculates TEMA alignment across 3 period sets (fast/standard/slow),
    adjusted by current volatility. Returns consensus signal and confidence.

    Returns:
        - consensus: -1.0 to +1.0 (signal strength)
        - confidence: 0.0 to 1.0 (agreement level)
        - ensemble_str: "X/3" format (e.g., "3/3", "2/3")
        - detail: Human-readable breakdown
    """
    # Get volatility scalar
    vol_scalar = get_volatility_scalar(close)

    # Base period sets (fast, mid, slow)
    base_periods = [
        (15, 40, 150),   # Fast: Catches early moves, more whipsaws
        (20, 50, 200),   # Standard: Industry standard
        (30, 75, 250),   # Slow: Filters noise, lags entries
    ]

    # Apply volatility adjustment
    adjusted_periods = [
        (int(fast * vol_scalar), int(mid * vol_scalar), int(slow * vol_scalar))
        for fast, mid, slow in base_periods
    ]

    alignment_signals = []
    details = []

    for (p_fast, p_mid, p_slow), (base_fast, base_mid, base_slow) in zip(adjusted_periods, base_periods):
        # Calculate TEMAs for this period set
        try:
            tema_fast = calculate_tema(close, p_fast).iloc[-1]
            tema_mid = calculate_tema(close, p_mid).iloc[-1]
            tema_slow = calculate_tema(close, p_slow).iloc[-1]
        except:
            # Skip this period set if calculation fails
            continue

        # Count bullish/bearish components
        score = 0
        if price > tema_fast:
            score += 1
        if tema_fast > tema_mid:
            score += 1
        if tema_mid > tema_slow:
            score += 1

        # Convert to signal: 3=bullish, 0=bearish, else=mixed
        if score == 3:
            signal = +1  # All bullish
        elif score == 0:
            signal = -1  # All bearish
        else:
            signal = 0   # Mixed

        alignment_signals.append(signal)

        # Detail string showing base periods and signal
        signal_str = "+" if signal > 0 else ("-" if signal < 0 else "~")
        details.append(f"{base_fast}/{base_mid}/{base_slow}:{signal_str}")

    # Calculate consensus and confidence
    if not alignment_signals:
        return 0.0, 0.0, "0/3", "Error", vol_scalar

    consensus = sum(alignment_signals) / len(alignment_signals)
    confidence = abs(consensus)

    # Count strong signals
    bullish_count = sum(1 for s in alignment_signals if s > 0)
    bearish_count = sum(1 for s in alignment_signals if s < 0)

    if consensus > 0:
        ensemble_str = f"{bullish_count}/3"
    elif consensus < 0:
        ensemble_str = f"-{bearish_count}/3"
    else:
        ensemble_str = "0/3"

    detail = " | ".join(details)

    return round(consensus, 2), round(confidence, 2), ensemble_str, detail, round(vol_scalar, 2)


def detect_cross(ma_fast, ma_fast_prev, ma_slow, ma_slow_prev):
    """
    Detect MA crossover.
    Returns: "Bullish Cross", "Bearish Cross", or "None"
    """
    if ma_fast > ma_slow and ma_fast_prev <= ma_slow_prev:
        return "Bullish Cross"
    elif ma_fast < ma_slow and ma_fast_prev >= ma_slow_prev:
        return "Bearish Cross"
    return "None"


def calculate_tsmom(close, lookbacks=[4, 12, 26]):
    """
    Time-series momentum: average return across lookback periods
    Returns the mean percentage return across all lookback windows
    """
    if len(close) < max(lookbacks):
        return None, []

    returns = []
    details = []
    for lb in lookbacks:
        ret = (close.iloc[-1] / close.iloc[-lb] - 1) * 100
        returns.append(ret)
        details.append(f"{lb}w: {ret:+.1f}%")

    composite = sum(returns) / len(returns)
    return round(composite, 2), details


def calculate_ma_score(close, price):
    """
    MA trend alignment score (0-7):
    - Price vs MA20, MA50, MA100, MA200
    - MA20 vs MA50, MA50 vs MA100, MA100 vs MA200
    """
    if len(close) < 50:
        return None, None, []

    ma20 = close.rolling(20).mean().iloc[-1]
    ma50 = close.rolling(50).mean().iloc[-1]
    ma100 = close.rolling(100).mean().iloc[-1] if len(close) >= 100 else None
    ma200 = close.rolling(200).mean().iloc[-1] if len(close) >= 200 else None

    checks = [
        (price > ma20, "Price>MA20"),
        (price > ma50, "Price>MA50"),
        (ma20 > ma50, "MA20>MA50"),
    ]

    if ma100 is not None:
        checks.extend([
            (price > ma100, "Price>MA100"),
            (ma50 > ma100, "MA50>MA100"),
        ])

    if ma200 is not None:
        checks.extend([
            (price > ma200, "Price>MA200"),
            (ma100 > ma200 if ma100 else ma50 > ma200, "MA100>MA200" if ma100 else "MA50>MA200"),
        ])

    score = sum(1 for cond, _ in checks if cond)
    max_score = len(checks)
    details = [name for cond, name in checks if cond]

    return score, max_score, details


def classify_regime(adx, tsmom_score, zscore, ma_score, ma_max):
    """
    Classify market regime for strategy selection.
    Returns: regime name, action bias

    tsmom_score is now the average % return across lookback periods
    """
    if adx is None or tsmom_score is None:
        return "UNKNOWN", "Insufficient data"

    ma_pct = (ma_score / ma_max) if ma_max > 0 else 0

    # Strong uptrend: high ADX + positive momentum + good MA alignment
    if adx > 25 and tsmom_score > 2 and ma_pct >= 0.6:
        return "TRENDING_UP", "Ride trend, buy dips"

    # Strong downtrend: high ADX + negative momentum
    if adx > 25 and tsmom_score < -2:
        return "TRENDING_DOWN", "Avoid or exit"

    # Mean-reversion opportunity: weak trend + extreme Z
    if adx < 25 and zscore is not None and abs(zscore) > 1.5:
        if zscore < -1.5:
            return "MEAN_REVERT_BUY", "Z-score oversold"
        else:
            return "MEAN_REVERT_SELL", "Z-score overbought"

    # Choppy/unclear
    if adx < 20:
        return "CHOPPY", "Reduce exposure, wait"

    return "NEUTRAL", "No strong edge"


def calculate_zscore(close, window=ZSCORE_WINDOW):
    """Calculate z-score for price series."""
    if len(close) < window:
        return None, None

    mean = float(close.rolling(window).mean().iloc[-1])
    std = float(close.rolling(window).std().iloc[-1])

    if std == 0 or np.isnan(std):
        return 0, "Neutral"

    zscore = round((float(close.iloc[-1]) - mean) / std, 2)

    # Classify zone using threshold ranges
    zones = [
        (2.5, float('inf'), "Extreme OB"),
        (2, 2.5, "Overbought"),
        (1, 2, "Upper"),
        (-1, 1, "Neutral"),
        (-2, -1, "Lower"),
        (-2.5, -2, "Oversold"),
        (float('-inf'), -2.5, "Extreme OS"),
    ]

    zone = next((z for low, high, z in zones if low <= zscore < high), "Neutral")
    return zscore, zone


def format_ma_distance(close, price, periods):
    """Calculate distance from MAs."""
    parts = []
    for p in periods:
        if len(close) >= p:
            ma = close.rolling(p).mean().iloc[-1]
            pct = ((price - ma) / ma) * 100
            parts.append(f"MA{p}: {abs(pct):.1f}%{'↑' if pct > 0 else '↓'}")
    return " | ".join(parts) if parts else "N/A"


def detect_trend(df):
    """Detect trend using MAs, ADX, and directional indicators."""
    if len(df) < 50:
        return "Insufficient Data", "N/A", None

    close, high, low = df['Close'], df['High'], df['Low']
    price = close.iloc[-1]

    ma20 = close.rolling(20).mean().iloc[-1]
    ma50 = close.rolling(50).mean().iloc[-1]
    ma100 = close.rolling(100).mean().iloc[-1] if len(close) >= 100 else None
    ma200 = close.rolling(200).mean().iloc[-1] if len(close) >= 200 else None

    adx_ind = ADXIndicator(high, low, close, window=14)
    adx = adx_ind.adx().iloc[-1]
    plus_di, minus_di = adx_ind.adx_pos().iloc[-1], adx_ind.adx_neg().iloc[-1]

    score = sum([
        1 if price > ma20 else -1,
        1 if price > ma50 else -1,
        1 if ma20 > ma50 else -1,
        1 if plus_di > minus_di else -1,
    ])
    if ma100 is not None:
        score += (1 if price > ma100 else -1) + (1 if ma50 > ma100 else -1)
    if ma200 is not None:
        ma_above_200 = ma100 > ma200 if ma100 else ma50 > ma200
        score += (1 if price > ma200 else -1) + (1 if ma_above_200 else -1)
    
    if adx < 20:
        return "↔️ Sideways/Choppy", "Weak", round(adx, 1)
    
    strength = "Moderate" if adx < 25 else "Strong"
    threshold = 3 if adx < 25 else 2
    
    if score >= threshold:
        trend = "📈 Uptrend"
    elif score <= -threshold:
        trend = "📉 Downtrend"
    else:
        trend = "↔️ Sideways/Choppy"
    
    return trend, strength, round(adx, 1)


def calculate_technicals(ticker):
    """Calculate technical indicators using weekly timeframe only."""
    # Download weekly data directly (complete weeks only)
    # Need ~260 weeks for MA200 weekly (5 years)
    weekly_df = get_cached_ticker(ticker, period="5y", interval="1wk")

    if weekly_df.empty or len(weekly_df) < 50:
        # Fallback: try shorter period
        weekly_df = get_cached_ticker(ticker, period="2y", interval="1wk")

    if weekly_df.empty:
        return None

    # Ensure we only use complete weekly candles (ending Sunday)
    # If last bar is not Sunday (weekday 6), drop it (incomplete week)
    last_bar_date = pd.Timestamp(weekly_df.index[-1])
    if last_bar_date.weekday() != 6:  # Not Sunday
        weekly_df = weekly_df[:-1]
        if weekly_df.empty:
            return None

    # Get price from last COMPLETE weekly candle close
    price = weekly_df['Close'].iloc[-1]
    weekly_date = pd.Timestamp(weekly_df.index[-1]).strftime('%Y-%m-%d')  # Store last complete week date

    weeks_available = len(weekly_df)

    if weeks_available < 26:
        return {
            'price': price,
            'weekly_date': weekly_date,
            'weeks': weeks_available,
            'zscore': None,
            'zscore_zone': 'N/A',
            'ma_distance': 'N/A',
            'trend': 'Insufficient Data',
            'trend_strength': 'N/A',
            'adx': None,
            'tsmom_score': None,
            'tsmom_details': [],
            'ma_score': None,
            'ma_max': None,
            'ma_details': [],
            'regime': 'UNKNOWN',
            'regime_bias': 'Insufficient data'
        }

    # Calculate indicators on weekly data
    close_weekly = weekly_df['Close']
    zscore, zone = calculate_zscore(close_weekly)
    ma_distance = format_ma_distance(close_weekly, price, MA_PERIODS)

    if len(weekly_df) >= 50:
        trend, trend_strength, adx = detect_trend(weekly_df)
    else:
        trend, trend_strength, adx = "Insufficient Data", "N/A", None

    # Trend-following indicators
    tsmom_score, tsmom_details = calculate_tsmom(close_weekly)
    ma_score, ma_max, ma_details = calculate_ma_score(close_weekly, price)

    # Regime classification
    regime, regime_bias = classify_regime(adx, tsmom_score, zscore, ma_score, ma_max)

    return {
        'price': price,
        'weekly_date': weekly_date,
        'weeks': weeks_available,
        'zscore': zscore,
        'zscore_zone': zone,
        'ma_distance': ma_distance,
        'trend': trend,
        'trend_strength': trend_strength,
        'adx': adx,
        'tsmom_score': tsmom_score,
        'tsmom_details': tsmom_details,
        'ma_score': ma_score,
        'ma_max': ma_max,
        'ma_details': ma_details,
        'regime': regime,
        'regime_bias': regime_bias
    }


def calculate_daily_technicals(ticker):
    """Calculate daily (1d) technical indicators using TEMA crosses."""
    # Fetch 1 year of daily data
    data = get_cached_ticker(ticker, period="1y", interval="1d")

    if data.empty or len(data) < 200:
        return None

    close = data['Close']
    high = data['High']
    low = data['Low']
    price = close.iloc[-1]
    daily_date = pd.Timestamp(data.index[-1]).strftime('%Y-%m-%d')  # Store last daily candle date

    # Daily Z-score (20-day window)
    zscore_daily, zone_daily = calculate_zscore(close, window=20)

    # Calculate TEMA for 20, 50, 200 periods
    tema20 = calculate_tema(close, 20)
    tema50 = calculate_tema(close, 50)
    tema200 = calculate_tema(close, 200)

    # Current and previous values for cross detection
    tema20_curr = tema20.iloc[-1]
    tema20_prev = tema20.iloc[-2]
    tema50_curr = tema50.iloc[-1]
    tema50_prev = tema50.iloc[-2]
    tema200_curr = tema200.iloc[-1]
    tema200_prev = tema200.iloc[-2]

    # Detect crosses
    cross_20_50 = detect_cross(tema20_curr, tema20_prev, tema50_curr, tema50_prev)
    cross_50_200 = detect_cross(tema50_curr, tema50_prev, tema200_curr, tema200_prev)

    # TEMA alignment (similar to MA score but with TEMA)
    tema_alignment = 0
    if tema20_curr > tema50_curr:
        tema_alignment += 1
    if tema50_curr > tema200_curr:
        tema_alignment += 1
    if price > tema20_curr:
        tema_alignment += 1

    # Daily ADX
    adx_ind = ADXIndicator(high, low, close, window=14)
    adx_daily = adx_ind.adx().iloc[-1]
    plus_di_daily = adx_ind.adx_pos().iloc[-1]
    minus_di_daily = adx_ind.adx_neg().iloc[-1]

    # Daily trend classification
    if tema20_curr > tema50_curr > tema200_curr and price > tema20_curr:
        trend_daily = "Strong Bullish"
    elif tema20_curr > tema50_curr and price > tema20_curr:
        trend_daily = "Bullish"
    elif tema20_curr < tema50_curr < tema200_curr and price < tema20_curr:
        trend_daily = "Strong Bearish"
    elif tema20_curr < tema50_curr and price < tema20_curr:
        trend_daily = "Bearish"
    else:
        trend_daily = "Mixed"

    # Distance from TEMAs (%)
    tema20_dist = ((price - tema20_curr) / tema20_curr) * 100
    tema50_dist = ((price - tema50_curr) / tema50_curr) * 100
    tema200_dist = ((price - tema200_curr) / tema200_curr) * 100

    # Multi-period TEMA ensemble with volatility adjustment
    consensus, confidence, ensemble_str, ensemble_detail, vol_scalar = calculate_tema_ensemble(close, price)

    # Enhanced trend classification using ensemble
    if confidence >= 0.67:  # At least 2/3 periods agree
        if consensus > 0:
            trend_ensemble = "Strong Bullish" if confidence == 1.0 else "Bullish"
        elif consensus < 0:
            trend_ensemble = "Strong Bearish" if confidence == 1.0 else "Bearish"
        else:
            trend_ensemble = "Mixed"
    else:
        trend_ensemble = "Low Confidence"

    return {
        'price': price,
        'daily_date': daily_date,
        'zscore_daily': zscore_daily,
        'zscore_zone_daily': zone_daily,
        'tema20': round(tema20_curr, 2),
        'tema50': round(tema50_curr, 2),
        'tema200': round(tema200_curr, 2),
        'tema20_dist': round(tema20_dist, 2),
        'tema50_dist': round(tema50_dist, 2),
        'tema200_dist': round(tema200_dist, 2),
        'cross_20_50': cross_20_50,
        'cross_50_200': cross_50_200,
        'tema_alignment': f"{tema_alignment}/3",
        'adx_daily': round(adx_daily, 1),
        'plus_di_daily': round(plus_di_daily, 1),
        'minus_di_daily': round(minus_di_daily, 1),
        'trend_daily': trend_daily,
        # NEW: Ensemble metrics
        'tema_consensus': consensus,
        'tema_confidence': confidence,
        'tema_ensemble': ensemble_str,
        'trend_ensemble': trend_ensemble,
        'vol_scalar': vol_scalar,
    }


# HELPERS

def get_regime_from_vix_z(vix_z):
    """Map VIX Z-score to market regime."""
    if vix_z >= 1.5:
        return "Complacency"
    elif vix_z <= -1.5:
        return "Fear"
    elif vix_z >= 0.5:
        return "Risk-On"
    elif vix_z <= -0.5:
        return "Risk-Off"
    return "Neutral"


def format_sign(value):
    """Format number with + or - sign."""
    return f"{'+' if value >= 0 else ''}{value}"


# CONDITIONAL FORMATTING

def optimize_column_widths(ws):
    """Auto-adjust column widths based on content."""
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                if cell.value:
                    # Convert to string and measure length
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass

        # Set width with some padding (add 2 for comfort)
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 to avoid excessive width
        ws.column_dimensions[column_letter].width = adjusted_width


def apply_conditional_formatting(xlsx_path, num_rows):
    """
    Apply conditional formatting and optimize column widths for Excel file.

    Color Scheme:
    - Z-scores: Red (extreme ±2) → White (neutral 0) → Yellow (moderate)
    - Returns/TSMOM: Red (negative) → White (0) → Green (positive)
    - ADX/Scores: White (low) → Green (high)
    - VIX: Green (low) → Red (high) - inverted scale

    Args:
        xlsx_path: Path to the Excel file
        num_rows: Number of data rows (excluding header) - used for dynamic range formatting
    """
    wb = load_workbook(xlsx_path)

    # === WEEKLY SHEET ===
    if 'Weekly' in wb.sheetnames:
        ws = wb['Weekly']

        # Z-Score: -2 (red) → 0 (white) → +2 (yellow/orange)
        ws.conditional_formatting.add(f'D2:D{num_rows+1}',
            ColorScaleRule(start_type='num', start_value=-2, start_color='F8696B',  # Red
                          mid_type='num', mid_value=0, mid_color='FFFFFF',          # White
                          end_type='num', end_value=2, end_color='FFEB84'))         # Yellow

        # TSMOM_%: -20 (red) → 0 (white) → +20 (green)
        ws.conditional_formatting.add(f'E2:E{num_rows+1}',
            ColorScaleRule(start_type='num', start_value=-20, start_color='F8696B',  # Red
                          mid_type='num', mid_value=0, mid_color='FFFFFF',           # White
                          end_type='num', end_value=20, end_color='63BE7B'))         # Green

        # MA_Score: 0 (white) → 7 (green)
        ws.conditional_formatting.add(f'F2:F{num_rows+1}',
            ColorScaleRule(start_type='num', start_value=0, start_color='FFFFFF',   # White
                          end_type='num', end_value=7, end_color='63BE7B'))          # Green

        # ADX: 10 (white) → 50 (green)
        ws.conditional_formatting.add(f'H2:H{num_rows+1}',
            ColorScaleRule(start_type='num', start_value=10, start_color='FFFFFF',  # White
                          end_type='num', end_value=50, end_color='63BE7B'))         # Green

        # Optimize column widths
        optimize_column_widths(ws)

    # === MOMENTUM SHEET ===
    if 'Momentum' in wb.sheetnames:
        ws = wb['Momentum']

        # All return columns: -30 (red) → 0 (white) → +30 (green)
        for col in ['C', 'D', 'E']:  # 4w, 12w, 26w returns
            ws.conditional_formatting.add(f'{col}2:{col}{num_rows+1}',
                ColorScaleRule(start_type='num', start_value=-30, start_color='F8696B',  # Red
                              mid_type='num', mid_value=0, mid_color='FFFFFF',           # White
                              end_type='num', end_value=30, end_color='63BE7B'))         # Green

        # Optimize column widths
        optimize_column_widths(ws)

    # === DAILY SHEET ===
    if 'Daily' in wb.sheetnames:
        ws = wb['Daily']

        # Z-Score: -2 (red) → 0 (white) → +2 (yellow)
        ws.conditional_formatting.add(f'D2:D{num_rows+1}',
            ColorScaleRule(start_type='num', start_value=-2, start_color='F8696B',
                          mid_type='num', mid_value=0, mid_color='FFFFFF',
                          end_type='num', end_value=2, end_color='FFEB84'))

        # ADX: 10 (white) → 50 (green)
        ws.conditional_formatting.add(f'G2:G{num_rows+1}',
            ColorScaleRule(start_type='num', start_value=10, start_color='FFFFFF',
                          end_type='num', end_value=50, end_color='63BE7B'))

        # Crosses: Highlight Bullish in green, Bearish in red
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

        # Use formula-based rules for text contains
        ws.conditional_formatting.add(f'F2:F{num_rows+1}',
            Rule(type='containsText', operator='containsText', formula=['SEARCH("Bullish",F2)'],
                 dxf=DifferentialStyle(fill=green_fill), text='Bullish'))
        ws.conditional_formatting.add(f'F2:F{num_rows+1}',
            Rule(type='containsText', operator='containsText', formula=['SEARCH("Bearish",F2)'],
                 dxf=DifferentialStyle(fill=red_fill), text='Bearish'))

        # Consensus: -1 (red) → 0 (white) → +1 (green)
        ws.conditional_formatting.add(f'H2:H{num_rows+1}',
            ColorScaleRule(start_type='num', start_value=-1, start_color='F8696B',
                          mid_type='num', mid_value=0, mid_color='FFFFFF',
                          end_type='num', end_value=1, end_color='63BE7B'))

        # Confidence: 0 (white) → 1 (green)
        ws.conditional_formatting.add(f'I2:I{num_rows+1}',
            ColorScaleRule(start_type='num', start_value=0, start_color='FFFFFF',
                          end_type='num', end_value=1, end_color='63BE7B'))

        # Optimize column widths
        optimize_column_widths(ws)

    # === MACRO SHEET ===
    if 'Macro' in wb.sheetnames:
        ws = wb['Macro']

        # VIX (row 2): 10 (green/calm) → 40 (red/fear) - INVERTED
        ws.conditional_formatting.add('B2:B2',
            ColorScaleRule(start_type='num', start_value=10, start_color='63BE7B',  # Green
                          end_type='num', end_value=40, end_color='F8696B'))        # Red

        # -Z(VIX) (row 3): -2 (red/fear) → 0 (white) → +2 (orange/complacency)
        ws.conditional_formatting.add('B3:B3',
            ColorScaleRule(start_type='num', start_value=-2, start_color='F8696B',
                          mid_type='num', mid_value=0, mid_color='FFFFFF',
                          end_type='num', end_value=2, end_color='FFEB84'))

        # F&G indices (rows 4-5): 0 (red/fear) → 50 (white) → 100 (red/greed)
        ws.conditional_formatting.add('B4:B5',
            ColorScaleRule(start_type='num', start_value=0, start_color='F8696B',
                          mid_type='num', mid_value=50, mid_color='FFFFFF',
                          end_type='num', end_value=100, end_color='F8696B'))

        # Optimize column widths
        optimize_column_widths(ws)

    wb.save(xlsx_path)
    print(f"  - Applied conditional formatting and optimized column widths")


# MAIN

def main():
    # Parse command line argument for assets file
    assets_file = sys.argv[1] if len(sys.argv) > 1 else None
    ASSETS = load_assets(assets_file)

    # Get assets file name for display
    assets_name = Path(assets_file).stem if assets_file else "assets"

    # Generate timestamp-based filename (24-hour format, no seconds)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M')
    output_dir = SCRIPT_DIR / "output"
    output_dir.mkdir(exist_ok=True)

    print(f"Fetching market data...")
    print(f"Date: {datetime.now().strftime('%A, %Y-%m-%d %H:%M')}")

    # Cleanup old orphan caches
    cleanup_orphan_caches(max_age_days=7)
    print(f"Portfolio: {assets_name}")

    # Collect macro data
    macro_data = []

    # GLI
    print("Fetching GLI data...")
    try:
        if gli := calculate_gli():
            macro_data.append({
                'Indicator': 'Global Liquidity',
                'Value': round(gli['value'], 2),
                'Unit': 'Billions USD',
                'Signal': gli['trend'],
                'Detail': f"4w: {format_sign(gli['mom_pct'])}% | 12w: {format_sign(gli['qoq_pct'])}%"
            })
        else:
            macro_data.append({'Indicator': 'Global Liquidity', 'Value': None, 'Unit': None, 'Signal': 'Error', 'Detail': 'Error'})
    except Exception:
        macro_data.append({'Indicator': 'Global Liquidity', 'Value': None, 'Unit': None, 'Signal': 'Error', 'Detail': 'Error'})

    # VIX
    print("Fetching VIX data...")
    try:
        vix, vix_z = get_vix_zscore()
        if vix:
            vix_levels = [(15, "Low"), (20, "Normal"), (30, "Elevated"), (float('inf'), "High")]
            vix_desc = next(desc for threshold, desc in vix_levels if vix < threshold)
            regime = get_regime_from_vix_z(vix_z)
            macro_data.append({'Indicator': 'VIX', 'Value': round(vix, 2), 'Unit': 'Index', 'Signal': vix_desc, 'Detail': ''})
            macro_data.append({'Indicator': '-Z(VIX)', 'Value': round(vix_z, 2), 'Unit': 'Z-Score', 'Signal': regime, 'Detail': ''})
    except Exception:
        macro_data.append({'Indicator': 'VIX', 'Value': None, 'Unit': None, 'Signal': 'Error', 'Detail': 'Error'})

    # Fear & Greed
    print("Fetching Fear & Greed indices...")
    for fn, lbl in [(get_fear_greed_traditional, 'F&G Stocks'), (get_fear_greed_crypto, 'F&G Crypto')]:
        try:
            val, cls = fn()
            macro_data.append({'Indicator': lbl, 'Value': int(val), 'Unit': '0-100 Scale', 'Signal': cls, 'Detail': ''})
        except Exception:
            macro_data.append({'Indicator': lbl, 'Value': None, 'Unit': None, 'Signal': 'Error', 'Detail': 'Error'})

    # Fetch all asset data (weekly and daily)
    print(f"Fetching data for {len(ASSETS)} assets...")
    asset_data = {}
    daily_data = {}
    for name, ticker in ASSETS.items():
        try:
            print(f"  - {ticker} (weekly + daily)...")
            asset_data[name] = calculate_technicals(ticker)
            daily_data[name] = calculate_daily_technicals(ticker)
        except Exception as e:
            print(f"    ERROR: {e}")
            asset_data[name] = None
            daily_data[name] = None

    # Extract dates from first available ticker data
    weekly_candle_date = None
    daily_candle_date = None
    for name in ASSETS.keys():
        if asset_data.get(name) and asset_data[name].get('weekly_date'):
            weekly_candle_date = asset_data[name]['weekly_date']
            break
    for name in ASSETS.keys():
        if daily_data.get(name) and daily_data[name].get('daily_date'):
            daily_candle_date = daily_data[name]['daily_date']
            break

    # Add data timeframe info to macro section
    if weekly_candle_date:
        macro_data.insert(0, {
            'Indicator': 'Weekly Data (Complete Week)',
            'Value': weekly_candle_date,
            'Unit': 'Date',
            'Signal': 'Last Complete',
            'Detail': 'All weekly indicators use this candle'
        })
    if daily_candle_date:
        macro_data.insert(1 if weekly_candle_date else 0, {
            'Indicator': 'Daily Data (Most Recent)',
            'Value': daily_candle_date,
            'Unit': 'Date',
            'Signal': 'Latest Available',
            'Detail': 'All daily indicators use this candle'
        })

    # Prepare asset analysis data
    asset_rows = []
    momentum_rows = []
    daily_rows = []

    for name, ticker in ASSETS.items():
        tech = asset_data.get(name)
        if tech:
            # Asset analysis row - use numeric types for Excel
            asset_rows.append({
                'Name': name,
                'Ticker': ticker,
                'Price': round(tech['price'], 4),
                'Z-Score': round(tech['zscore'], 2) if tech['zscore'] is not None else None,
                'TSMOM_%': round(tech['tsmom_score'], 2) if tech['tsmom_score'] is not None else None,
                'MA_Score': tech['ma_score'] if tech['ma_score'] is not None else None,
                'MA_Max': tech['ma_max'] if tech['ma_max'] is not None else None,
                'ADX': round(tech['adx'], 1) if tech['adx'] is not None else None,
                'Regime': tech['regime'],
                'Regime_Bias': tech['regime_bias']
            })

            # Momentum details row - extract numeric values
            if tech.get('tsmom_details'):
                details = tech['tsmom_details']
                # Parse "4w: +2.5%" -> 2.5
                ret_4w = float(details[0].split(': ')[1].rstrip('%')) if len(details) > 0 else None
                ret_12w = float(details[1].split(': ')[1].rstrip('%')) if len(details) > 1 else None
                ret_26w = float(details[2].split(': ')[1].rstrip('%')) if len(details) > 2 else None

                momentum_rows.append({
                    'Name': name,
                    'Ticker': ticker,
                    '4w_Return_%': ret_4w,
                    '12w_Return_%': ret_12w,
                    '26w_Return_%': ret_26w,
                    'MA_Distance': tech['ma_distance']  # Keep as text (complex format)
                })
        else:
            asset_rows.append({
                'Name': name,
                'Ticker': ticker,
                'Price': None,
                'Z-Score': None,
                'TSMOM_%': None,
                'MA_Score': None,
                'MA_Max': None,
                'ADX': None,
                'Regime': 'Error',
                'Regime_Bias': 'Error'
            })

        # Daily technicals row - use numeric types for Excel
        daily_tech = daily_data.get(name)
        if daily_tech:
            # Consolidate TEMA distances into one readable string
            tema_dist = f"20:{daily_tech['tema20_dist']:+.1f}% | 50:{daily_tech['tema50_dist']:+.1f}% | 200:{daily_tech['tema200_dist']:+.1f}%"

            # Consolidate crosses into one column
            crosses = []
            if daily_tech['cross_20_50'] != 'None':
                crosses.append(f"20x50:{daily_tech['cross_20_50'].replace(' Cross', '')}")
            if daily_tech['cross_50_200'] != 'None':
                crosses.append(f"50x200:{daily_tech['cross_50_200'].replace(' Cross', '')}")
            crosses_str = " | ".join(crosses) if crosses else "None"

            daily_rows.append({
                'Name': name,
                'Ticker': ticker,
                'Price': round(daily_tech['price'], 4),
                'Z-Score': round(daily_tech['zscore_daily'], 2) if daily_tech['zscore_daily'] is not None else None,
                'TEMA_Dist': tema_dist,
                'Crosses': crosses_str,
                'ADX': round(daily_tech['adx_daily'], 1),
                'Consensus': round(daily_tech['tema_consensus'], 2),
                'Confidence': round(daily_tech['tema_confidence'], 2),
                'Ensemble': daily_tech['tema_ensemble'],
                'Trend': daily_tech['trend_ensemble'],
                'Vol_Scalar': round(daily_tech['vol_scalar'], 2),
            })
        else:
            daily_rows.append({
                'Name': name,
                'Ticker': ticker,
                'Price': None,
                'Z-Score': None,
                'TEMA_Dist': 'Error',
                'Crosses': 'Error',
                'ADX': None,
                'Consensus': None,
                'Confidence': None,
                'Ensemble': 'Error',
                'Trend': 'Error',
                'Vol_Scalar': None,
            })

    # Write XLSX file with multiple sheets
    print("\nWriting XLSX file...")

    xlsx_file = output_dir / f"{timestamp}_ANALYSIS.xlsx"

    with pd.ExcelWriter(xlsx_file, engine='openpyxl') as writer:
        # Sheet 1: Macro indicators
        if macro_data:
            df_macro = pd.DataFrame(macro_data)
            df_macro.to_excel(writer, sheet_name='Macro', index=False)

        # Sheet 2: Weekly signals
        if asset_rows:
            df_weekly = pd.DataFrame(asset_rows)
            df_weekly.to_excel(writer, sheet_name='Weekly', index=False)

        # Sheet 3: Momentum details
        if momentum_rows:
            df_momentum = pd.DataFrame(momentum_rows)
            df_momentum.to_excel(writer, sheet_name='Momentum', index=False)

        # Sheet 4: Daily signals
        if daily_rows:
            df_daily = pd.DataFrame(daily_rows)
            df_daily.to_excel(writer, sheet_name='Daily', index=False)

    print(f"  - {xlsx_file}")

    # Apply conditional formatting
    print("\nApplying conditional formatting...")
    apply_conditional_formatting(xlsx_file, len(asset_rows))
    print("  - Formatting applied")

    print(f"\nAnalysis complete. File saved to: {output_dir}")


if __name__ == "__main__":
    main()
