"""Excel file generation and conditional formatting."""

from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule, Rule
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.styles.differential import DifferentialStyle
from reporting.formatters import optimize_column_widths


def apply_borders(ws):
    """Apply thin borders to all cells with data in the worksheet."""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border


def apply_conditional_formatting(xlsx_path, num_rows):
    """
    Apply conditional formatting and optimize column widths for Excel file.

    Color Scheme:
    - ZTanh: Red (extreme -1) → White (neutral 0) → Yellow (extreme +1)
    - Returns/TSMOM: Red (negative) → White (0) → Green (positive)
    - ADX/Scores: White (low) → Green (high)
    - VIX: Green (low) → Red (high) - inverted scale

    Args:
        xlsx_path: Path to the Excel file
        num_rows: Number of data rows (excluding header) - used for dynamic range formatting
    """
    wb = load_workbook(xlsx_path)

    # === WEEKLY SHEET ===
    # Columns: Name, Ticker, Price, ZTanh, ZTanh_Zone, TSMOM_%, MA_Score, MA_Max, ADX, Regime, Regime_Bias
    if 'Weekly' in wb.sheetnames:
        ws = wb['Weekly']

        # ZTanh (col D): -1 (red) → 0 (white) → +1 (yellow/orange)
        ws.conditional_formatting.add(f'D2:D{num_rows+1}',
            ColorScaleRule(start_type='num', start_value=-1, start_color='F8696B',  # Red
                          mid_type='num', mid_value=0, mid_color='FFFFFF',          # White
                          end_type='num', end_value=1, end_color='FFEB84'))         # Yellow

        # TSMOM_% (col F): -20 (red) → 0 (white) → +20 (green)
        ws.conditional_formatting.add(f'F2:F{num_rows+1}',
            ColorScaleRule(start_type='num', start_value=-20, start_color='F8696B',  # Red
                          mid_type='num', mid_value=0, mid_color='FFFFFF',           # White
                          end_type='num', end_value=20, end_color='63BE7B'))         # Green

        # MA_Score (col G): 0 (white) → 7 (green)
        ws.conditional_formatting.add(f'G2:G{num_rows+1}',
            ColorScaleRule(start_type='num', start_value=0, start_color='FFFFFF',   # White
                          end_type='num', end_value=7, end_color='63BE7B'))          # Green

        # ADX (col I): 10 (white) → 50 (green)
        ws.conditional_formatting.add(f'I2:I{num_rows+1}',
            ColorScaleRule(start_type='num', start_value=10, start_color='FFFFFF',  # White
                          end_type='num', end_value=50, end_color='63BE7B'))         # Green

        # Apply borders and optimize column widths
        apply_borders(ws)
        optimize_column_widths(ws)

    # === MOMENTUM SHEET ===
    # Columns: Name, Ticker, 4w_Return_%, 12w_Return_%, 26w_Return_%, MA_Distance
    if 'Momentum' in wb.sheetnames:
        ws = wb['Momentum']

        # All return columns (C, D, E): -30 (red) → 0 (white) → +30 (green)
        for col in ['C', 'D', 'E']:  # 4w, 12w, 26w returns
            ws.conditional_formatting.add(f'{col}2:{col}{num_rows+1}',
                ColorScaleRule(start_type='num', start_value=-30, start_color='F8696B',  # Red
                              mid_type='num', mid_value=0, mid_color='FFFFFF',           # White
                              end_type='num', end_value=30, end_color='63BE7B'))         # Green

        # Apply borders and optimize column widths
        apply_borders(ws)
        optimize_column_widths(ws)

    # === DAILY SHEET ===
    # Columns: Name, Ticker, Price, ZTanh, ZTanh_Zone, MA_Score, MA_Dist, ADX, ADX_Action, DI_Bias, Trend
    if 'Daily' in wb.sheetnames:
        ws = wb['Daily']

        # ZTanh (col D): -1 (red) → 0 (white) → +1 (yellow)
        ws.conditional_formatting.add(f'D2:D{num_rows+1}',
            ColorScaleRule(start_type='num', start_value=-1, start_color='F8696B',
                          mid_type='num', mid_value=0, mid_color='FFFFFF',
                          end_type='num', end_value=1, end_color='FFEB84'))

        # MA_Score (col F): 0 (white) → 7 (green)
        ws.conditional_formatting.add(f'F2:F{num_rows+1}',
            ColorScaleRule(start_type='num', start_value=0, start_color='FFFFFF',
                          end_type='num', end_value=7, end_color='63BE7B'))

        # ADX (col H): 10 (white) → 50 (green)
        ws.conditional_formatting.add(f'H2:H{num_rows+1}',
            ColorScaleRule(start_type='num', start_value=10, start_color='FFFFFF',
                          end_type='num', end_value=50, end_color='63BE7B'))

        # DI_Bias (col J): Highlight Bullish in green, Bearish in red
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

        ws.conditional_formatting.add(f'J2:J{num_rows+1}',
            Rule(type='containsText', operator='containsText', formula=['SEARCH("Bullish",J2)'],
                 dxf=DifferentialStyle(fill=green_fill), text='Bullish'))
        ws.conditional_formatting.add(f'J2:J{num_rows+1}',
            Rule(type='containsText', operator='containsText', formula=['SEARCH("Bearish",J2)'],
                 dxf=DifferentialStyle(fill=red_fill), text='Bearish'))

        # Trend (col K): Highlight based on trend direction
        ws.conditional_formatting.add(f'K2:K{num_rows+1}',
            Rule(type='containsText', operator='containsText', formula=['SEARCH("Bullish",K2)'],
                 dxf=DifferentialStyle(fill=green_fill), text='Bullish'))
        ws.conditional_formatting.add(f'K2:K{num_rows+1}',
            Rule(type='containsText', operator='containsText', formula=['SEARCH("Bearish",K2)'],
                 dxf=DifferentialStyle(fill=red_fill), text='Bearish'))

        # Apply borders and optimize column widths
        apply_borders(ws)
        optimize_column_widths(ws)

    # === MACRO SHEET ===
    if 'Macro' in wb.sheetnames:
        ws = wb['Macro']

        # VIX: 10 (green/calm) → 40 (red/fear) - INVERTED
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=15, min_col=1, max_col=1), start=1):
            if row[0].value == 'VIX':
                ws.conditional_formatting.add(f'B{row_idx}:B{row_idx}',
                    ColorScaleRule(start_type='num', start_value=10, start_color='63BE7B',  # Green
                                  end_type='num', end_value=40, end_color='F8696B'))        # Red

        # VIX ZTanh: -1 (green/fear=opportunity) → 0 (white) → +1 (red/greed=caution)
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=15, min_col=1, max_col=1), start=1):
            if row[0].value == 'VIX ZTanh':
                ws.conditional_formatting.add(f'B{row_idx}:B{row_idx}',
                    ColorScaleRule(start_type='num', start_value=-1, start_color='63BE7B',  # Green (fear)
                                  mid_type='num', mid_value=0, mid_color='FFFFFF',          # White
                                  end_type='num', end_value=1, end_color='F8696B'))         # Red (greed)

        # Apply borders and optimize column widths
        apply_borders(ws)
        optimize_column_widths(ws)

    wb.save(xlsx_path)
    print(f"  - Applied conditional formatting and optimized column widths")
